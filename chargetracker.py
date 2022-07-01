# Script skapad av LEM
from datetime import datetime
from openpyxl.workbook.workbook import Workbook
import requests
import time
import sys
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
# Statuskoder i api.chargefinder.com
# 2 = Ledig (GRÖN)
# 3 = Upptagen (BLÅ)
# 5 = Otillgänglig (OKÄND)
ROW_RANGE = 3

def init_workbook():
	try:
		wb = openpyxl.load_workbook("chargers_data.xlsx")
	except:
		wb = Workbook()
	return wb

def find_empty_grids(ws):
	start_col = ws.max_column + 2
	row_col = [1, start_col]
	return row_col

def insert_date(row_col, ws, data):
	now = datetime.now()
	index = 0
	ledig_total = 0
	upptagen_total = 0
	otillgänglig_total = 0
	dt_string = now.strftime("%Y/%m/%d %H:%M:%S")
	ws.cell(row=row_col[0], column=row_col[1]).value = dt_string
	ws.merge_cells(start_row=row_col[0], start_column=row_col[1], end_row=row_col[0], end_column=row_col[1]+ROW_RANGE)
	ws.cell(row=row_col[0]+1, column=row_col[1]).value = "PLATS"
	ws.cell(row=row_col[0]+1, column=row_col[1]+1).value = "LEDIG"
	ws.cell(row=row_col[0]+1, column=row_col[1]+2).value = "UPPTAGEN"
	ws.cell(row=row_col[0]+1, column=row_col[1]+3).value = "OTILLGÄNGLIG"
	for i, item in enumerate(data):
		if item["Status"]:
			ws.cell(row=row_col[0]+2+index, column=row_col[1]).value = item["Namn"]
			ws.cell(row=row_col[0]+2+index, column=row_col[1]+1).value = item["Ledig"]
			ws.cell(row=row_col[0]+2+index, column=row_col[1]+2).value = item["Upptagen"]
			ws.cell(row=row_col[0]+2+index, column=row_col[1]+3).value = item["Otillgänglig"]
			index = index + 1
			ledig_total = ledig_total + item["Ledig"]
			upptagen_total = upptagen_total + item["Upptagen"]
			otillgänglig_total = otillgänglig_total + item["Otillgänglig"]
	# Skriv summa
	ws.cell(row=row_col[0]+2+index+1, column=row_col[1]+1).value = ledig_total
	ws.cell(row=row_col[0]+2+index+1, column=row_col[1]+2).value = upptagen_total
	ws.cell(row=row_col[0]+2+index+1, column=row_col[1]+3).value = otillgänglig_total

	dim_holder = DimensionHolder(worksheet=ws)
	for col in range(ws.min_column, ws.max_column + 1):
		dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
	ws.column_dimensions = dim_holder

def store_data(arguments, wb, result):
	arguments_formatted = [x.replace("-", " ") for x in arguments]
	worksheet_title = " -> ".join(arguments_formatted)
	if worksheet_title in wb.sheetnames:
		ws = wb[worksheet_title]
		row_col = find_empty_grids(ws)
		insert_date(row_col, ws, result)
	else:
		ws = wb.create_sheet(worksheet_title)
		row_col=[1,1]
		insert_date(row_col, ws, result)

# Hämta start, slut
def load_start_end(arguments):
	towns = []
	for x in arguments:
		metadata = {}
		place = x.replace("-", "%20")
		rer = requests.get('https://photon.komoot.io/api/?limit=10&q={}'.format(place))
		metadata["input_place"] = rer.json()['features'][0]['properties']['type']
		metadata["longitude"] = rer.json()['features'][0]['geometry']['coordinates'][0]
		metadata["latitude"] = rer.json()['features'][0]['geometry']['coordinates'][1]
		towns.append(metadata)
	return towns

wb = init_workbook()

towns = load_start_end(sys.argv[1:])
r = requests.get('https://api.chargefinder.com/route?from={}&fromlat={}&fromlng={}&fromcc=&via=&vialat=&vialng=&to={}&tolat={}&tolng={}&preference=recommended&detour=4&minspeed=3&maxspeed=6'.format(
	towns[0]["input_place"],
	towns[0]["latitude"],
	towns[0]["longitude"],
	towns[1]["input_place"],
	towns[1]["latitude"],
	towns[1]["longitude"],
	)
	)
data = r.json()

result = []
counter = 1
now = datetime.now()
dt_string = now.strftime("%d/%m/%Y %H:%M:%S")

for station in data['stations']:
	station_data = {}
	station_id = station['slug']
	station_req = requests.get('https://api.chargefinder.com/station/{}'.format(station_id))
	station_status_req = requests.get('https://api.chargefinder.com/status/{}'.format(station_id))
	if station_status_req.text != "null" and station_status_req.status_code != 502:
		station_data['Namn'] = station_req.json()['title']
		station_data['Status'] = True
		station_data['Ledig'] = sum([1 for d in station_status_req.json() if d['status'] == 2])
		station_data['Upptagen'] = sum([1 for d in station_status_req.json() if d['status'] == 3])
		station_data['Otillgänglig'] = sum([1 for d in station_status_req.json() if d['status'] == 5])
	else:
		station_data['Status'] = False
	result.append(station_data)
	print("Processing..... {} / {}".format(counter, len(data['stations'])))
	counter = counter + 1

print("Scan genomförd: {}".format(dt_string))
store_data(sys.argv[1:], wb, result)
wb.save("chargers_data.xlsx")
